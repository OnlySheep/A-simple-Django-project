from django.core.paginator import Paginator, EmptyPage
from django.shortcuts import render, HttpResponse, redirect
from .models import Student, StudentDetail, Clas, Course, UserInfo
import os
from openpyxl import load_workbook
from django.contrib import auth


# 认证用户装饰器
# def outer(origin):
#     def inner(request, *args, **kwargs):
#         if not request.user.username:
#             return redirect("/login/")
#         res = origin(request, *args, **kwargs)
#         return res
#
#     return inner


# Create your views here.

def index(request):
    # 分页对象
    student_list = Student.objects.all()
    paginator = Paginator(student_list, 8)

    # 分页信息
    print(paginator.count)  # 11
    print(paginator.num_pages)  # 2
    print(paginator.page_range)  # 分页列表  range(1, 3)

    try:
        # （2）获取某页对象
        page = int(request.GET.get("page", 1))
        student_list = paginator.page(page)
        # 获取该页数据
        # 方式1：
        print(student_list)
        print(student_list.object_list)

        # 方式2：
        for student in student_list:
            print(student)
    except EmptyPage:
        student_list = paginator.page(1)
    # for student in students:
    #     courses = student.courses.all()
    #     course_list = list(course.title for course in courses)
    #     total_list.append(
    #         [[student.name], [student.age], ["男" if student.sex == 1 else "女"], [student.birthday.strftime("%F")],
    #          course_list])
    # print(total_list)
    title = "学生信息"
    return render(request, "student/index.html", {"student_list": student_list, "title": title,
                                                  "page_range": paginator.page_range, "page_num": page,
                                                  "num_pages": paginator.num_pages})


def add_student(request):
    name = request.POST.get("name")
    age = request.POST.get("age")
    sex = request.POST.get("sex")
    birthday = request.POST.get("birthday")
    clas_id = request.POST.get("clas_id")
    if request.method == "GET":
        clas_list = Clas.objects.all()
        title = "学生添加"
        return render(request, "student/add.html", {"clas_list": clas_list, "title": title})
    else:
        if name and age and sex and birthday and clas_id:
            student_request = request.POST
            print(student_request)
            name = request.POST.get("name")
            if not Student.objects.filter(name=name):
                student_dict = student_request.dict()
                del student_dict["csrfmiddlewaretoken"]  # student_dict.pop("csrfmiddlewaretoken")也可以
                Student.objects.create(**student_dict)
                return redirect("/student/")
            else:
                tips = "该学生已存在！"
                clas_list = Clas.objects.all()
                title = "学生添加"
                return render(request, "student/add.html", {"clas_list": clas_list, "tips": tips, "title": title})
        else:
            tips = "内容不能为空！"
            clas_list = Clas.objects.all()
            title = "学生添加"
            return render(request, "student/add.html", {"clas_list": clas_list, "tips": tips, "title": title})


def delete_student(request, del_id):
    student = Student.objects.get(pk=del_id)
    if request.method == "GET":
        title = "删除学生"
        return render(request, "student/delete.html", {"student": student, "title": title})

    else:
        student.delete()  # 级联删除的原因
        return redirect("/student/")


def update_student(request, upd_id):
    student = Student.objects.get(pk=upd_id)
    clas_list = Clas.objects.all()
    course_list = Course.objects.all()
    if request.method == "GET":
        title = "修改学生"
        return render(request, "student/update.html",
                      {"student": student, "clas_list": clas_list, "course_list": course_list, "title": title})
    else:
        name = request.POST.get("name")
        age = request.POST.get("age")
        sex = request.POST.get("sex")
        birthday = request.POST.get("birthday")
        clas_id = request.POST.get("clas_id")
        print(request.POST)
        if name and age and sex and birthday and clas_id:
            course_id_list = request.POST.getlist("course_id_list")
            data = request.POST.dict()
            data.popitem()  # 删除字典最后一个，或着用data.pop("course_id_list")
            data.pop("csrfmiddlewaretoken")
            Student.objects.filter(pk=upd_id).update(**data)
            student.courses.set(course_id_list)
            return redirect("/student/")
        else:
            tips = "内容不能为空！"
            title = "修改学生"
            clas_list = Clas.objects.all()
            course_list = Course.objects.all()
            return render(request, "student/update.html",
                          {"tips": tips, "title": title, "student": student, "clas_list": clas_list,
                           "course_list": course_list})


def select_student(request):
    # print(request.GET)
    # content = request.GET.get("content")
    # if not content:
    #     tips = "查询内容不能为空哦！"
    #     total_list = Student.objects.all()
    #     title = "首页"
    #     return render(request, "student/index.html",
    #                   {"tips": tips, "total_list": total_list, "title": title, "content": content})
    # way = request.GET.get("select_way")
    # if way == "follow_name":
    #     student_list = Student.objects.filter(name__contains=content)
    #     if not student_list:
    #         tips = "没有该学生哦!"
    #         total_list = Student.objects.all()
    #         title = "首页"
    #         return render(request, "student/index.html",
    #                       {"tips": tips, "total_list": total_list, "title": title, "content": content})
    #     else:
    #         title = "查询学生"
    #     return render(request, "student/select.html",
    #                   {"student_list": student_list, "title": title, "content": content})
    # elif way == "follow_class":
    #     clas = Clas.objects.filter(name=content)
    #     if not clas:
    #         tips = "没有该班级哦!"
    #         total_list = Student.objects.all()
    #         title = "首页"
    #         return render(request, "student/index.html",
    #                       {"tips": tips, "total_list": total_list, "title": title, "content": content})
    #     student_list = clas[0].student_list.all()  # 一对多或者多对多记得用all取
    #     title = "查询学生"
    #     return render(request, "student/select.html",
    #                   {"student_list": student_list, "title": title, "content": content})
    select_way = request.GET.get("select_way")
    key_word = request.GET.get("key_word")
    if not key_word:
        return redirect("/")

    if select_way == "follow_name":

        student_list = Student.objects.filter(name__contains=key_word)

    elif select_way == "follow_class":
        student_list = Student.objects.filter(clas__name=key_word)
    else:
        student_list = []

    return render(request, "student/index.html", {"student_list": student_list, "key_word": key_word})


def course_index(request):
    # 分页对象
    course_list = Course.objects.all()
    paginator = Paginator(course_list, 8)

    # 分页信息
    print(paginator.count)  # 11
    print(paginator.num_pages)  # 2
    print(paginator.page_range)  # 分页列表  range(1, 3)

    try:
        # （2）获取某页对象
        page = int(request.GET.get("page", 1))
        course_list = paginator.page(page)
        # 获取该页数据
        # 方式1：
        print(course_list)
        print(course_list.object_list)

        # 方式2：
        for student in course_list:
            print(student)
    except EmptyPage:
        course_list = paginator.page(1)
    title = "课程信息"
    return render(request, "student/course_index.html", {"course_list": course_list, "title": title,
                                                         "page_range": paginator.page_range, "page_num": page,
                                                         "num_pages": paginator.num_pages})


def add_course(request):
    if request.method == "GET":
        title = "添加课程"
        return render(request, "student/add2.html", {"title": title})
    else:
        title = request.POST.get("title")
        credit = request.POST.get("credit")
        teacher = request.POST.get("teacher")
        course_request = request.POST
        if title and credit and teacher:
            if not Course.objects.filter(title=title):
                course_dict = course_request.dict()
                print(course_dict)
                del course_dict["csrfmiddlewaretoken"]  # course_dict.pop("csrfmiddlewaretoken")也可以
                Course.objects.create(**course_dict)
                return redirect("/student/course_index/")
            else:
                tips = "该课程已存在！"
                title = "课程添加"
                return render(request, "student/add2.html", {"tips": tips, "title": title})
        # course = request.POST.get("course")  # 不能用成了 request.POST.course
        # Course.objects.create(title=course)
        else:
            tips = "课程名,学分或讲师内容不能为空！"
            title = "课程添加"
            return render(request, "student/add2.html", {"tips": tips, "title": title})


def delete_course(request, del_id):
    course = Course.objects.get(pk=del_id)
    if request.method == "GET":
        title = "删除课程"
        return render(request, "student/delete2.html", {"course": course, "title": title})
    else:
        course.delete()  # 级联删除的原因
        return redirect("/student/course_index/")


def update_course(request, upd_id):
    course = Course.objects.get(pk=upd_id)
    if request.method == "GET":
        title = "编辑课程"
        return render(request, "student/update2.html", {"course": course, "title": title})
    else:
        title = request.POST.get("title")
        credit = request.POST.get("credit")
        teacher = request.POST.get("teacher")
        data = request.POST.dict()
        data.pop("csrfmiddlewaretoken")
        if title and credit and teacher:
            Course.objects.filter(pk=upd_id).update(**data)
            return redirect("/student/course_index/")
        else:
            tips = "课程名,学分或讲师内容不能为空！"
            title = "编辑课程"
            return render(request, "student/update2.html", {"tips": tips, "title": title, "course": course})


def select_course(request):
    select_way = request.GET.get("select_way")
    key_word = request.GET.get("key_word")
    if not key_word:  # 判断key_word是否为空
        return redirect("/student/course_index/")

    if select_way == "follow_title":
        course_list = Course.objects.filter(title__contains=key_word)
    elif select_way == "follow_teacher":
        course_list = Course.objects.filter(teacher__contains=key_word)
    else:
        course_list = []

    return render(request, "student/course_index.html", {"course_list": course_list, "key_word": key_word})


def stu_excel(request):
    # （1）保存表格
    excel_file = request.FILES.get("excel_file")
    print(excel_file)
    path = os.path.join("media", "files", excel_file.name)
    with open(path, "wb") as f:
        for line in excel_file:
            f.write(line)

    # (2)通过python操作excel表格
    wb = load_workbook(path)
    print(wb.sheetnames)
    work_sheet = wb.worksheets[0]
    student_list = list()
    for line in work_sheet.iter_rows(min_row=2):
        # print("line:", line)
        # for cell in line:
        #     print("cell:", cell.value)
        if not line[0].value:
            break
        # 一对一：
        stu = StudentDetail.objects.create(tel=line[4].value, addr=line[5].value)
        sex = 1 if line[2].value == "男" else 0
        clas_id = Clas.objects.get(name=line[6].value).id
        # 先实例化对象
        student = Student(name=line[0].value,
                          age=line[1].value,
                          sex=sex,
                          birthday=line[3].value,
                          clas_id=clas_id,
                          stu_detail=stu)
        student_list.append(student)
    # 再添加
    Student.objects.bulk_create(student_list)
    return redirect("/student/")


def elective(request):
    if request.method == "GET":
        # 分页对象
        course_list = Course.objects.all()
        paginator = Paginator(course_list, 8)

        # 分页信息
        print(paginator.count)  # 11
        print(paginator.num_pages)  # 2
        print(paginator.page_range)  # 分页列表  range(1, 3)

        try:
            # （2）获取某页对象
            page = int(request.GET.get("page", 1))
            course_list = paginator.page(page)
            # 获取该页数据
            # 方式1：
            print(course_list)
            print(course_list.object_list)

            # 方式2：
            for student in course_list:
                print(student)
        except EmptyPage:
            course_list = paginator.page(1)
        title = "选择课程"
        return render(request, "student/course.html", {"course_list": course_list, "title": title,
                                                       "page_range": paginator.page_range, "page_num": page,
                                                       "num_pages": paginator.num_pages})
    else:
        print(request.POST.getlist("course_id_list"))
        course_id_list = request.POST.getlist("course_id_list")
        student_id = request.user.stu_id
        student = Student.objects.get(pk=student_id)
        student.courses.set(course_id_list)
        return redirect("/")


def add_clas(request):
    if request.method == "GET":
        return render(request, "student/add3.html")
    else:
        name = request.POST.get("name")
        if name:
            course = Clas.objects.filter(name=name)
            if not course:
                Clas.objects.create(name=name)
                return redirect("/")
            else:
                tips = "该班级已经存在"
                return render(request, "student/add3.html", {"tips": tips, "name": name})
        else:
            tips = "内容不能为空"
            return render(request, "student/add3.html", {"tips": tips, "name": name})


def logout(request):
    auth.logout(request)
    return redirect("/login/")


def login(request):
    if request.method == "GET":
        return render(request, "student/login.html")
    else:
        username = request.POST.get("username")
        pwd = request.POST.get("pwd")

        # 用户认证组件
        user = auth.authenticate(username=username, password=pwd)

        if user:
            # 认证成功
            # 写session:request.session["user_id"] = user_id
            auth.login(request, user)
            return redirect("/")
        else:
            # 认证失败
            tips = "用户名或密码输入错误"
            return render(request, "student/login.html", {"tips": tips})
